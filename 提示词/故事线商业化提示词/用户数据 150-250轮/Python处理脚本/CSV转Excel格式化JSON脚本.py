#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV转Excel格式化JSON脚本
用途：将conversation_summaries.csv转换为Excel格式，并格式化JSON内容
"""

import pandas as pd
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

class CSVToExcelFormatter:
    def __init__(self, csv_file_path: str):
        """
        初始化转换器
        
        Args:
            csv_file_path: CSV文件路径
        """
        self.csv_file_path = csv_file_path
        self.df = None
        
    def load_csv_data(self) -> bool:
        """
        加载CSV数据
        
        Returns:
            是否成功加载数据
        """
        try:
            print("📖 正在读取CSV文件...")
            # 读取CSV文件，指定编码为utf-8，添加正确的列名
            column_names = ['ROUND_COUNT', 'SESSION_ID', 'SEGMENT_NO', 'SESSION_SUMMARY', 'START_TIME', 'END_TIME', 'DIALOGUE_SEGMENTS']
            self.df = pd.read_csv(self.csv_file_path, encoding='utf-8', names=column_names, header=None)
            print(f"✅ 成功读取CSV文件，共{len(self.df)}行数据")
            
            # 打印列名
            print(f"📋 列名: {list(self.df.columns)}")
            return True
            
        except Exception as e:
            print(f"❌ 读取CSV文件失败: {str(e)}")
            return False
    
    def format_json_column(self, column_name: str) -> None:
        """
        格式化指定列的JSON内容
        
        Args:
            column_name: 需要格式化的列名
        """
        if column_name not in self.df.columns:
            print(f"⚠️ 列 '{column_name}' 不存在，跳过格式化")
            return
        
        print(f"🔧 正在格式化 {column_name} 列的JSON内容...")
        formatted_count = 0
        error_count = 0
        
        for index, row in self.df.iterrows():
            try:
                json_str = row[column_name]
                if pd.isna(json_str) or json_str == '':
                    continue
                    
                # 尝试解析JSON
                if isinstance(json_str, str):
                    json_obj = json.loads(json_str)
                    # 格式化JSON，增加缩进和确保中文正常显示
                    formatted_json = json.dumps(
                        json_obj, 
                        ensure_ascii=False, 
                        indent=2,
                        separators=(',', ': ')
                    )
                    self.df.at[index, column_name] = formatted_json
                    formatted_count += 1
                    
            except json.JSONDecodeError as e:
                error_count += 1
                print(f"⚠️ 第{index+1}行JSON解析错误: {str(e)}")
                # 保留原始内容不变
                continue
            except Exception as e:
                error_count += 1
                print(f"⚠️ 第{index+1}行处理错误: {str(e)}")
                continue
        
        print(f"✅ {column_name} 列格式化完成: 成功{formatted_count}个，错误{error_count}个")
    
    def create_excel_with_formatting(self, output_path: str) -> bool:
        """
        创建格式化的Excel文件
        
        Args:
            output_path: 输出Excel文件路径
            
        Returns:
            是否成功创建文件
        """
        try:
            print("📊 正在创建Excel文件...")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Conversation Summaries"
            
            # 添加数据到工作表
            for r in dataframe_to_rows(self.df, index=False, header=True):
                ws.append(r)
            
            # 设置样式
            self._apply_excel_formatting(ws)
            
            # 保存文件
            wb.save(output_path)
            print(f"✅ Excel文件已保存到: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ 创建Excel文件失败: {str(e)}")
            return False
    
    def _apply_excel_formatting(self, worksheet):
        """
        应用Excel格式化
        
        Args:
            worksheet: Excel工作表对象
        """
        # 定义样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 获取数据范围
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        print(f"🎨 正在应用格式化样式 ({max_row}行 x {max_col}列)...")
        
        # 应用标题行格式
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 应用数据行格式
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = border
        
        # 自动调整列宽
        self._auto_adjust_column_width(worksheet)
        
        # 冻结首行
        worksheet.freeze_panes = 'A2'
        
        print("✅ 格式化样式应用完成")
    
    def _auto_adjust_column_width(self, worksheet):
        """
        自动调整列宽
        
        Args:
            worksheet: Excel工作表对象
        """
        print("📏 正在自动调整列宽...")
        
        column_widths = {}
        
        # 计算每列的最佳宽度
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    column = cell.column_letter
                    # 对于JSON列，设置固定宽度以便查看
                    if isinstance(cell.value, str) and ('{' in cell.value or '[' in cell.value):
                        column_widths[column] = max(column_widths.get(column, 0), 80)
                    else:
                        # 其他列根据内容长度调整
                        content_length = len(str(cell.value))
                        column_widths[column] = max(column_widths.get(column, 0), min(content_length + 2, 50))
        
        # 应用列宽
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        print("✅ 列宽调整完成")
    
    def process_file(self, output_filename: str = None) -> str:
        """
        处理整个文件转换流程
        
        Args:
            output_filename: 输出文件名，默认自动生成
            
        Returns:
            输出文件路径
        """
        # 1. 加载CSV数据
        if not self.load_csv_data():
            return None
        
        # 2. 格式化JSON列（根据实际列名调整）
        json_columns = []
        
        # 检测可能的JSON列
        for col in self.df.columns:
            # 检查列中是否包含JSON格式的内容
            sample_values = self.df[col].dropna().head(3)
            for value in sample_values:
                if isinstance(value, str) and (value.strip().startswith('{') or value.strip().startswith('[')):
                    json_columns.append(col)
                    break
        
        print(f"🔍 检测到JSON列: {json_columns}")
        
        # 格式化所有JSON列
        for col in json_columns:
            self.format_json_column(col)
        
        # 3. 生成输出文件路径
        if not output_filename:
            base_name = os.path.splitext(os.path.basename(self.csv_file_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{base_name}_格式化_{timestamp}.xlsx"
        
        output_dir = os.path.dirname(self.csv_file_path)
        output_path = os.path.join(output_dir, output_filename)
        
        # 4. 创建Excel文件
        if self.create_excel_with_formatting(output_path):
            return output_path
        else:
            return None


def main():
    """主函数"""
    # CSV文件路径
    csv_file_path = "/Users/edy/Desktop/project/挑战玩法/提示词/故事线商业化提示词/用户数据 150-250轮/conversation_summaries.csv"
    
    # 检查文件是否存在
    if not os.path.exists(csv_file_path):
        print(f"❌ CSV文件不存在: {csv_file_path}")
        return
    
    print("="*60)
    print("📊 CSV转Excel格式化工具")
    print("="*60)
    print(f"输入文件: {os.path.basename(csv_file_path)}")
    
    # 创建转换器并处理文件
    formatter = CSVToExcelFormatter(csv_file_path)
    output_path = formatter.process_file("conversation_summaries_格式化.xlsx")
    
    if output_path:
        print("="*60)
        print("🎉 转换完成!")
        print(f"📁 输出文件: {os.path.basename(output_path)}")
        print(f"📂 完整路径: {output_path}")
        print("="*60)
        
        # 显示文件统计信息
        if formatter.df is not None:
            print("📊 文件统计:")
            print(f"  总行数: {len(formatter.df)}")
            print(f"  总列数: {len(formatter.df.columns)}")
            print(f"  文件大小: {os.path.getsize(output_path) / (1024*1024):.2f} MB")
    else:
        print("❌ 转换失败!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n❌ 用户中断了程序")
    except Exception as e:
        print(f"❌ 程序执行出错: {str(e)}")
        import traceback
        traceback.print_exc()