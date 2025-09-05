#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel中JSON列格式化脚本
读取Excel文件，格式化JSON列，输出美化后的Excel文件
"""

import pandas as pd
import json
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def format_json_in_excel(input_file: str, output_file: str):
    """
    格式化Excel文件中的JSON列
    
    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
    """
    print(f"📖 读取文件: {input_file}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(input_file)
        print(f"✅ 成功读取，共 {len(df)} 行数据")
        print(f"📊 列名: {list(df.columns)}")
        
        # 检查哪些列包含JSON数据
        json_columns = []
        for col in df.columns:
            # 检查第一个非空值是否是JSON格式
            for val in df[col].dropna():
                if isinstance(val, str) and (val.strip().startswith('{') or val.strip().startswith('[')):
                    json_columns.append(col)
                    break
        
        print(f"🔍 检测到JSON列: {json_columns}")
        
        if not json_columns:
            print("❌ 未检测到JSON列")
            return
        
        # 格式化JSON列
        for col in json_columns:
            print(f"🔧 格式化列: {col}")
            
            def format_json_cell(cell_value):
                if pd.isna(cell_value):
                    return cell_value
                
                try:
                    if isinstance(cell_value, str):
                        # 尝试解析JSON
                        json_obj = json.loads(cell_value)
                        # 美化格式
                        return json.dumps(json_obj, ensure_ascii=False, indent=2)
                    return cell_value
                except (json.JSONDecodeError, TypeError):
                    return cell_value
            
            df[col] = df[col].apply(format_json_cell)
        
        # 创建新的Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "格式化数据"
        
        # 添加数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置样式
        header_font = Font(bold=True, size=12)
        content_font = Font(size=10)
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # 格式化标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 格式化数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row):
                cell.font = content_font
                cell.alignment = wrap_alignment
                
                # 如果是JSON列，设置更大的行高和列宽
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                col_name = df.columns[col_idx]
                
                if col_name in json_columns:
                    ws.column_dimensions[col_letter].width = 80
                    ws.row_dimensions[row_idx].height = 200
        
        # 保存文件
        wb.save(output_file)
        print(f"💾 格式化完成，已保存到: {output_file}")
        
        # 显示统计信息
        print("\n📈 处理统计:")
        print(f"总行数: {len(df)}")
        print(f"总列数: {len(df.columns)}")
        print(f"JSON列数: {len(json_columns)}")
        
        return True
        
    except Exception as e:
        print(f"❌ 处理失败: {str(e)}")
        return False


def preview_json_columns(input_file: str, max_rows: int = 3):
    """
    预览Excel文件中的JSON列
    
    Args:
        input_file: 输入Excel文件路径
        max_rows: 预览的最大行数
    """
    try:
        df = pd.read_excel(input_file)
        print(f"📖 文件预览: {input_file}")
        print(f"📊 数据形状: {df.shape}")
        print(f"📋 列名: {list(df.columns)}")
        print("-" * 60)
        
        for col in df.columns:
            print(f"\n🔍 列: {col}")
            # 显示前几行的数据
            for i in range(min(max_rows, len(df))):
                val = df[col].iloc[i]
                if pd.isna(val):
                    print(f"  行{i+1}: [空值]")
                elif isinstance(val, str) and len(val) > 100:
                    print(f"  行{i+1}: {val[:100]}...")
                else:
                    print(f"  行{i+1}: {val}")
                    
    except Exception as e:
        print(f"❌ 预览失败: {str(e)}")


def main():
    """主函数"""
    base_dir = "/Users/edy/Desktop/project/挑战玩法/提示词/故事线商业化提示词/用户数据 150-250轮/评测完成"
    input_file = os.path.join(base_dir, "conversation_summaries.xlsx")
    output_file = os.path.join(base_dir, "conversation_summaries_格式化.xlsx")
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"❌ 文件不存在: {input_file}")
        return
    
    print("🚀 开始处理Excel文件JSON格式化")
    print("=" * 60)
    
    # 先预览文件结构
    preview_json_columns(input_file)
    
    print("\n" + "=" * 60)
    print("🔧 开始格式化处理...")
    
    # 执行格式化
    success = format_json_in_excel(input_file, output_file)
    
    if success:
        print(f"\n🎉 处理完成!")
        print(f"📁 原文件: {input_file}")
        print(f"📁 新文件: {output_file}")
    else:
        print(f"\n❌ 处理失败")


if __name__ == "__main__":
    main()